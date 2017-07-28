from openpyxl import load_workbook
import os
import re

rename_dir = ""
file_output_format = ".csv"

def createAuthorList(excel_file_path):
  """ (str) -> list of str
  
  Return a list of author id as authorList.
  
  >>> createAuthorList("C:\")
  ['1111111', '2222222']
  
  """
  wb = load_workbook(filename = excel_file_path)
  print(wb.get_sheet_names())
  
  #selects sheet1 and each row for column A
  ws = wb["Sheet1"]
  sheet_range = ws["A1:A" + str(ws.max_row)]
  authorList = []
  for cell in sheet_range:
    authorList.append(str(cell[0].value))
  return authorList

def rename(path):
  """ (str) -> None
  
  Renames .csv or .xls scopus author files in the directory path.
  
  >>> rename("C:\")
  
  """
  os.chdir(path) #opens this directory
  regex = r"[a-zA-Z. ']+"
  for file in os.listdir(os.getcwd()):
    try:
      if file.endswith(".csv") or file.endswith(".xls"):
        csv = open(file)
        line = csv.readline()
        matches = re.findall(regex, line)
        print(matches)
        csv.close()
        os.rename(file, matches[1] + matches[2] + file_output_format)
    except:
      continue
  print("RENAMING COMPLETE")

def authors(excel_file_path):
  """ (str) -> dict of str : list of str
  Returns a dictionary of last + first + affiliation keys from authorId.
  """
  wb = load_workbook(filename = excel_file_path)
  ws = wb["Sheet1"]
  names = {}
  i = 1
  while i < ws.max_row + 1:
    print(ws["B" + str(i)])
    last = ws["B" + str(i)].value.lower()
    first = ws["C" + str(i)].value.lower()
    if ws["D" + str(i)].value is None:
      affiliation = ' '
    else:
      affiliation = ws["D" + str(i)].value.lower()
    names[last + first + affiliation] = [last, first, affiliation]
    i += 1
  return names

def add_id(dict_name, excel_file_path):
  """ (dict of str:list of str, str) -> None
  
  Adds scopus author id to the authorId.xlsx 
  """
  wb = load_workbook(filename = excel_file_path)
  ws = wb["Sheet1"]
  i = 1
  while i < ws.max_row + 1:
    last = ws["B" + str(i)].value.lower()
    first = ws["C" + str(i)].value.lower()    
    if ws["D" + str(i)].value is None:
      affiliation = ' '
    else:
      affiliation = ws["D" + str(i)].value.lower()
    key = last + first + affiliation
    try:
      ws["A" + str(i)] = dict_name[key][3]
      ws["E" + str(i)] = dict_name[key][4]
    except:
      i += 1
      continue
    i += 1
  wb.save(excel_file_path)

